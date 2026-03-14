import pandas as pd
import os
import json
import logging
import asyncio
from datetime import datetime
from typing import List, Dict, Any, Optional, Union
from pydantic import BaseModel, Field
from langchain_core.tools import tool
from src.config import settings

logger = logging.getLogger(__name__)

# --- SCHEMAS EXPLICITOS ---

class ReadExcelSchema(BaseModel):
    file_path: Optional[str] = Field(default=None, description="O nome do arquivo Excel (ex: 'dados.xlsx') a ser lido.")
    sheet_name: Union[str, int, None] = Field(default=0, description="Opcional. Nome da aba ou índice (0 para a primeira).")

class WriteExcelSchema(BaseModel):
    data: Optional[Any] = Field(default=None, description="OBRIGATÓRIO. Os dados para criar a planilha.")
    file_path: Optional[str] = Field(default=None, description="OBRIGATÓRIO. O nome do arquivo Excel (ex: 'relatorio.xlsx').")
    sheet_name: Optional[str] = Field(default="Sheet1", description="Opcional. Nome da aba.")

def _clean_data(raw_data: Any) -> List[Dict[str, Any]]:
    """Tenta limpar e consertar o payload de dados enviado pelo modelo Gemini/OpenAI."""
    import re
    logger.info(f"📊 [ExcelTool] Recebido para limpeza: tipo={type(raw_data)}")
    
    if raw_data is None:
        logger.error("[ExcelTool] Erro: Dados recebidos como None.")
        raise ValueError("O campo 'data' foi recebido como None/Empty. Você deve prover dados para criar uma planilha.")
        
    if isinstance(raw_data, str):
        # Limpeza de markdown code blocks (```json ... ```)
        clean_str = re.sub(r'```(?:json)?\n?(.*?)\n?```', r'\1', raw_data, flags=re.DOTALL).strip()
        try:
            raw_data = json.loads(clean_str)
        except Exception:
            # Tenta extrair apenas a parte que parece JSON (entre [] ou {})
            match = re.search(r'(\[.*\]|\{.*\})', clean_str, re.DOTALL)
            if match:
                try:
                    raw_data = json.loads(match.group(1))
                except:
                    raise ValueError("O campo 'data' foi enviado como uma string mas não conseguimos extrair um JSON válido.")
            else:
                raise ValueError("O campo 'data' foi enviado como uma string mas não é um JSON válido.")
            
    if not isinstance(raw_data, list):
        if isinstance(raw_data, dict):
            raw_data = [raw_data]
        else:
            raise ValueError(f"O campo 'data' deve ser uma lista (array) de dicionários (objetos). Recebeu: {type(raw_data)}")
            
    cleaned = []
    for item in raw_data:
        if isinstance(item, dict):
            cleaned.append(item)
        elif isinstance(item, list):
            # Converte lista em colunas numeradas
            cleaned.append({f"Coluna_{i+1}": val for i, val in enumerate(item)})
        else:
            cleaned.append({"Valor": str(item)})
            
    if not cleaned:
        raise ValueError("A lista de dados estava vazia após a limpeza.")
        
    return cleaned

def _apply_premium_style(file_path: str):
    """Aplica o design corporativo premium (Inspirado no Manus AI) à planilha gerada."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = load_workbook(file_path)
        ws = wb.active
        
        # --- DEFINIÇÃO DE ESTILOS PREMIUM ---
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        entrada_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        saida_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        entrada_font = Font(color="006100")
        saida_font = Font(color="9C0006")
        
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB")
        )
        align_center = Alignment(horizontal="center", vertical="center")

        # --- CABEÇALHOS ---
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border

        # --- DADOS E FORMATAÇÃO CONDICIONAL ---
        valor_col_idx = -1
        tipo_col_idx = -1
        status_col_idx = -1

        # Identifica colunas-chave
        for col in range(1, ws.max_column + 1):
            header_val = str(ws.cell(row=1, column=col).value).lower()
            if "valor" in header_val: valor_col_idx = col
            if "tipo" in header_val: tipo_col_idx = col
            if "status" in header_val: status_col_idx = col

        # Status cores
        status_fills = {
            "recebido": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "pago": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
            "pendente": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "agendado": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        }

        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 25
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = align_center
                
                # Formata Moeda
                if col == valor_col_idx:
                    cell.number_format = '#,##0.00'

                # Cores por Tipo (Entrada/Saída)
                if col == tipo_col_idx:
                    val_tipo = str(cell.value).lower()
                    if "entrada" in val_tipo or "receita" in val_tipo:
                        cell.fill = entrada_fill
                        cell.font = entrada_font
                    else:
                        cell.fill = saida_fill
                        cell.font = saida_font
                
                # Cores por Status
                if col == status_col_idx:
                    val_status = str(cell.value).lower()
                    if val_status in status_fills:
                        cell.fill = status_fills[val_status]

        # --- RODAPÉ DE TOTAIS COM FÓRMULAS DINÂMICAS ---
        if valor_col_idx != -1 and ws.max_row > 1:
            start_row = ws.max_row + 2
            v_letter = get_column_letter(valor_col_idx)
            t_letter = get_column_letter(tipo_col_idx) if tipo_col_idx != -1 else None
            
            # TOTAL GERAL (Soma simples como fallback se não houver coluna de tipo)
            l1 = ws.cell(row=start_row, column=valor_col_idx-1, value="TOTAL GERAL:")
            l1.font = Font(bold=True)
            l1.alignment = Alignment(horizontal="right")
            
            v1 = ws.cell(row=start_row, column=valor_col_idx)
            v1.value = f"=SUM({v_letter}2:{v_letter}{ws.max_row})"
            v1.font = Font(bold=True)
            v1.number_format = '#,##0.00'
            v1.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            v1.border = thin_border

            # Se houver coluna de tipo, faz o Saldo (Entradas - Saídas)
            if t_letter:
                l2 = ws.cell(row=start_row+1, column=valor_col_idx-1, value="SALDO ESTIMADO:")
                l2.font = Font(bold=True, size=12)
                l2.alignment = Alignment(horizontal="right")
                
                v2 = ws.cell(row=start_row+1, column=valor_col_idx)
                # Fórmula SUMIF: Soma se tipo for entrada, subtrai se for saída
                v2.value = f'=SUMIF({t_letter}2:{t_letter}{ws.max_row}, "*receita*", {v_letter}2:{v_letter}{ws.max_row}) + SUMIF({t_letter}2:{t_letter}{ws.max_row}, "*entrada*", {v_letter}2:{v_letter}{ws.max_row}) - SUMIF({t_letter}2:{t_letter}{ws.max_row}, "*saída*", {v_letter}2:{v_letter}{ws.max_row}) - SUMIF({t_letter}2:{t_letter}{ws.max_row}, "*despesa*", {v_letter}2:{v_letter}{ws.max_row})'
                v2.font = Font(bold=True, size=12)
                v2.number_format = '#,##0.00'
                v2.border = thin_border

        # Ajuste de largura inteligente baseado no conteúdo
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 4)
            ws.column_dimensions[column].width = min(adjusted_width, 50)

        wb.save(file_path)
        logger.info(f"✨ [ExcelGen] Formatação Avançada aplicada em: {file_path}")
    except Exception as e:
        logger.warning(f"⚠️ [ExcelGen] Erro ao aplicar estilo avançado: {e}")

@tool(args_schema=ReadExcelSchema)
async def read_excel(file_path: str, sheet_name: Union[str, int] = 0) -> List[Dict[str, Any]]:
    """
    Lê uma planilha Excel e retorna os dados como uma lista de dicionários.
    Ideal para recuperar dados previamente gerados ou analisar planilhas existentes.
    """
    if file_path is None:
        logger.error("[ExcelTool] Erro crítico: file_path é None!")
        return [{"error": "file_path não pode ser nulo"}]
        
    try:
        if not os.path.isabs(file_path):
            file_path = os.path.join(settings.DATA_OUTPUTS_PATH, file_path)
            
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
            
        df = await asyncio.to_thread(pd.read_excel, file_path, sheet_name=sheet_name)
        data = df.to_dict(orient="records")
        logger.info(f"Sucesso ao ler Excel: {file_path} ({len(data)} linhas)")
        return data
    except Exception as e:
        logger.error(f"Erro ao ler Excel {file_path}: {e}")
        return [{"error": str(e)}]

@tool(args_schema=WriteExcelSchema)
async def create_excel(data: list, file_path: str, sheet_name: str = "Sheet1") -> str:
    """
    Cria UMA NOVA planilha Excel (.xlsx) inteira a partir de uma lista de dados JSON.
    Esta ferramenta deve ser usada para consolidar relatórios, tabelas de tendências ou novos documentos.
    O campo 'data' deve ser OBRIGATORIAMENTE uma lista (array).
    """
    if data is None or file_path is None:
        logger.error(f"[ExcelGen] ERRO CRÍTICO: Parâmetros nulos! data={data}, file_path={file_path}")
        if data is None: data = []
        if file_path is None: file_path = f"relatorio_{int(datetime.now().timestamp())}.xlsx"

    try:
        logger.info(f"🚀 [ExcelGen] Iniciando criação de {file_path}...")
        clean_data = _clean_data(data)
        
        if not file_path.endswith(".xlsx"):
            file_path += ".xlsx"
            
        full_path = os.path.join(settings.DATA_OUTPUTS_PATH, file_path)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        
        # --- TENTATIVA 1: PANDAS ---
        try:
            logger.info("[ExcelGen] Tentando via PANDAS...")
            df = pd.DataFrame(clean_data)
            await asyncio.to_thread(df.to_excel, full_path, index=False, sheet_name=sheet_name)
            await asyncio.to_thread(_apply_premium_style, full_path)
            logger.info(f"✅ [ExcelGen] Sucesso via PANDAS: {full_path}")
            size = os.path.getsize(full_path)
            return f"Planilha '{file_path}' gerada com sucesso ({size} bytes). <SEND_FILE:{os.path.basename(full_path)}>"
        except Exception as pe:
            logger.warning(f"⚠️ [ExcelGen] Pandas falhou: {pe}. Tentando FALLBACK OpenPyXL...")
            
            # --- TENTATIVA 2: OPENPYXL NATIVO ---
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            if clean_data:
                headers = list(clean_data[0].keys())
                ws.append(headers)
                for row in clean_data:
                    ws.append([row.get(h, "") for h in headers])
            
            await asyncio.to_thread(wb.save, full_path)
            await asyncio.to_thread(_apply_premium_style, full_path)
            logger.info(f"✅ [ExcelGen] Sucesso via FALLBACK OpenPyXL: {full_path}")
            return f"Planilha '{file_path}' criada via fallback seguro com {len(clean_data)} linhas. <SEND_FILE:{os.path.basename(full_path)}>"
            
    except Exception as e:
        logger.error(f"❌ [ExcelGen] Erro crítico ao criar Excel {file_path}: {e}")
        return f"Erro ao criar planilha: {str(e)}"

@tool(args_schema=WriteExcelSchema)
async def append_to_excel(data: list, file_path: str, sheet_name: str = "Sheet1") -> str:
    """
    Adiciona NOVAS LINHAS a uma planilha Excel já existente de forma incremental.
    """
    if data is None:
        logger.error("[ExcelTool] Erro: append_to_excel recebeu data=None. Usando lista vazia.")
        data = []
        
    try:
        clean_data = _clean_data(data)
        
        if not file_path.endswith(".xlsx"):
            file_path += ".xlsx"
            
        if not os.path.isabs(file_path):
            full_path = os.path.join(settings.DATA_OUTPUTS_PATH, file_path)
        else:
            full_path = file_path
            
        if not os.path.exists(full_path):
            return await create_excel.ainvoke({"data": clean_data, "file_path": file_path, "sheet_name": sheet_name})
            
        existing_df = await asyncio.to_thread(pd.read_excel, full_path, sheet_name=sheet_name)
        new_df = pd.DataFrame(clean_data)
        
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        await asyncio.to_thread(combined_df.to_excel, full_path, index=False, sheet_name=sheet_name)
        
        logger.info(f"Dados adicionados ao Excel: {full_path}")
        return f"Mais {len(clean_data)} linhas adicionadas à planilha '{os.path.basename(full_path)}'. <SEND_FILE:{os.path.basename(full_path)}>"
    except Exception as e:
        logger.error(f"Erro ao atualizar Excel {file_path}: {e}")
        return f"Erro ao atualizar planilha: {str(e)}"
