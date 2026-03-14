from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def gerar_excel_exemplo(nome_arquivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Controle Financeiro"

    # Estilos
    font_header = Font(color="FFFFFF", bold=True)
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Cabeçalhos
    headers = ["Data", "Descrição", "Tipo", "Valor", "Status"]
    ws.append(headers)
    for col_num, cell in enumerate(ws[1]):
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 15 # Largura padrão

    # Dados de exemplo
    data = [
        ("2026-03-01", "Salário", "Entrada", 3000.00, "Pago"),
        ("2026-03-05", "Aluguel", "Saída", 1200.00, "Pago"),
        ("2026-03-10", "Supermercado", "Saída", 350.75, "Pago"),
        ("2026-03-15", "Freelance", "Entrada", 800.00, "Pendente"),
        ("2026-03-20", "Conta de Luz", "Saída", 150.00, "Agendado"),
        ("2026-03-25", "Internet", "Saída", 99.90, "Pago"),
    ]

    # Adicionar dados e formatação condicional
    for row_data in data:
        ws.append(row_data)
        row_num = ws.max_row
        # Formatação de tipo (Entrada/Saída)
        if row_data[2] == "Entrada":
            ws[f'D{row_num}'].font = Font(color="008000") # Verde
        elif row_data[2] == "Saída":
            ws[f'D{row_num}'].font = Font(color="FF0000") # Vermelho
        
        # Formatação de status
        if row_data[4] == "Pago":
            ws[f'E{row_num}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Azul claro
        elif row_data[4] == "Pendente":
            ws[f'E{row_num}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarelo claro
        elif row_data[4] == "Agendado":
            ws[f'E{row_num}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Verde claro

    # Totais
    ws.append([]) # Linha em branco
    ws.append(["", "", "Total Entradas:", "=SUMIF(C:C,\"Entrada\",D:D)", ""])
    ws.append(["", "", "Total Saídas:", "=SUMIF(C:C,\"Saída\",D:D)", ""])
    ws.append(["", "", "Saldo Final:", "=D" + str(ws.max_row-1) + "+D" + str(ws.max_row), ""])

    # Ajustar largura das colunas automaticamente para o conteúdo
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(nome_arquivo)
    print(f'Planilha {nome_arquivo} gerada com sucesso.')

if __name__ == '__main__':
    gerar_excel_exemplo('controle_financeiro.xlsx')
