"""
GERADOR DE EXCEL PROFISSIONAL
Usa: openpyxl com formatação avançada
"""

import os
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def gerar_excel(titulo: str, cabecalhos: list, dados: list, nome_aba: str = "Dados") -> str:
    """
    Gera Excel profissional com formatação completa.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nome_aba

    azul = "1E3A8A"
    azul_medio = "3B82F6"
    branco = "FFFFFF"
    cinza = "F1F5F9"
    azul_claro = "DBEAFE"

    borda = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    # ── Título ──
    ws.merge_cells(f"A1:{get_column_letter(len(cabecalhos))}1")
    cell_titulo = ws["A1"]
    cell_titulo.value = titulo.upper()
    cell_titulo.font = Font(bold=True, size=16, color=branco, name="Segoe UI")
    cell_titulo.fill = PatternFill(start_color=azul, end_color=azul, fill_type="solid")
    cell_titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # ── Cabeçalhos ──
    for col, header in enumerate(cabecalhos, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color=branco, size=11, name="Segoe UI")
        cell.fill = PatternFill(start_color=azul_medio, end_color=azul_medio, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda
    ws.row_dimensions[2].height = 25

    # ── Dados ──
    for row_idx, row_data in enumerate(dados, 3):
        fill_cor = cinza if row_idx % 2 == 0 else branco
        for col_idx, valor in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill = PatternFill(start_color=fill_cor, end_color=fill_cor, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = borda
            cell.font = Font(name="Calibri", size=11)
            
            # Detectar e formatar números
            if isinstance(valor, float):
                cell.number_format = '#,##0.00'
            elif isinstance(valor, int):
                cell.number_format = '#,##0'
        ws.row_dimensions[row_idx].height = 22

    # ── Linha de Total (se houver dados numéricos) ──
    total_row = len(dados) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=11, name="Segoe UI", color=branco)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color=azul, end_color=azul, fill_type="solid")
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_row, column=1).border = borda

    for col_idx in range(2, len(cabecalhos) + 1):
        valores_col = []
        for row_data in dados:
            if col_idx - 1 < len(row_data) and isinstance(row_data[col_idx - 1], (int, float)):
                valores_col.append(row_data[col_idx - 1])
        
        cell = ws.cell(row=total_row, column=col_idx)
        if valores_col:
            cell.value = sum(valores_col)
            cell.font = Font(bold=True, size=11, name="Segoe UI", color=branco)
            cell.number_format = '#,##0.00' if any(isinstance(v, float) for v in valores_col) else '#,##0'
        cell.fill = PatternFill(start_color=azul, end_color=azul, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda
    ws.row_dimensions[total_row].height = 28

    # ── Ajustar largura das colunas automaticamente ──
    for col_idx, header in enumerate(cabecalhos, 1):
        max_len = len(str(header))
        for row_data in dados:
            if col_idx - 1 < len(row_data):
                max_len = max(max_len, len(str(row_data[col_idx - 1])))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    # ── Congelar painel no cabeçalho ──
    ws.freeze_panes = "A3"

    # Salvar
    os.makedirs("outputs", exist_ok=True)
    uid = str(uuid.uuid4())[:8]
    caminho = f"outputs/{titulo.replace(' ', '_')}_{uid}.xlsx"
    wb.save(caminho)
    print(f"✅ Excel gerado: {caminho}")
    return caminho
